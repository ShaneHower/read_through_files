import re
import docx2txt
import openpyxl
import os

# grab files in directory 
check_files = []
for dir, subdir, files in os.walk(r'C:\path\to\folder'):
    for file in files:
        check_files.append(file)

#activate excel workbook
wb = openpyxl.Workbook()
ws = wb.active

#run through each file and search it for item 
file_number = 1
for file in check_files:
    text = docx2txt.process(r'C:\path\to\folder\{0}'.format(file))

    # takes the whole line after Small Order Charge
    sm_ord_chrg = re.compile(r'SMALL ORDER CHARGE\s+.*')
   
   # puts it into a list
    mo_sm_ord_chrg = sm_ord_chrg.findall(text)
    
    #append it to the excel work book 
    ws.append(mo_sm_ord_chrg)

    # if the small order charge was found grad the invoice number from the file.  They happend to all start with 3 
    # and were 7 digits long 
    if len(mo_sm_ord_chrg) > 0:
        # grabs invoice number
        invoice_num = re.compile(r"3\d\d\d\d\d\d") 

        mo_invoice = invoice_num.findall(text)
        
        #upload inv number to the second column 
        ws.cell(row= file_number, column=2, value=mo_invoice[0])
    
    # update the console so we can see how far along the script is 
    file_number += 1
    print "number of files left: {0}".format(len(check_files)-file_number)

# make an id column so that we know that it searched every document 
id_count = 1
for i in range(len(check_files)):
    ws.cell(row=id_count, column=3, value=id_count)
    id_count += 1

wb.save("desired_name.xlsx")
